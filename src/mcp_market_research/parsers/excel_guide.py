from __future__ import annotations

import threading
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook import Workbook

from ..errors import ChannelNotFoundError, ConfigurationError, ExcelGuideValidationError
from ..models import ChannelGuide, ChannelSection, GuidanceRow, QuestionType

EXPECTED_COLUMNS = ("section", "question_type", "guidance", "example", "required", "notes")
REQUIRED_COLUMNS = ("section", "question_type", "guidance")
TRUTHY = frozenset({"yes", "y", "true", "1", "t"})


_cache_lock = threading.Lock()
_cache: dict[Path, tuple[float, dict[str, ChannelGuide]]] = {}


def list_channel_names(path: Path) -> list[str]:
    return list(load_channel_guides(path).keys())


def get_channel_guide(path: Path, channel: str) -> ChannelGuide:
    guides = load_channel_guides(path)
    normalized = channel.strip()
    if normalized in guides:
        return guides[normalized]
    lower_lookup = {k.lower(): k for k in guides}
    if normalized.lower() in lower_lookup:
        return guides[lower_lookup[normalized.lower()]]
    raise ChannelNotFoundError(channel, list(guides.keys()))


def load_channel_guides(path: Path) -> dict[str, ChannelGuide]:
    resolved = Path(path).resolve()
    if not resolved.exists():
        raise ConfigurationError(f"Excel guide not found at {resolved}")

    mtime = resolved.stat().st_mtime
    with _cache_lock:
        cached = _cache.get(resolved)
        if cached is not None and cached[0] == mtime:
            return cached[1]

    workbook = load_workbook(filename=resolved, read_only=True, data_only=True)
    try:
        guides = _parse_workbook(workbook)
    finally:
        workbook.close()

    with _cache_lock:
        _cache[resolved] = (mtime, guides)
    return guides


def clear_cache() -> None:
    with _cache_lock:
        _cache.clear()


def _parse_workbook(workbook: Workbook) -> dict[str, ChannelGuide]:
    guides: dict[str, ChannelGuide] = {}
    errors: list[str] = []

    for sheet_name in workbook.sheetnames:
        if sheet_name.startswith("_"):
            continue
        try:
            guides[sheet_name] = _parse_sheet(workbook[sheet_name], sheet_name)
        except _SheetValidationError as exc:
            errors.extend(exc.messages)

    if errors:
        raise ExcelGuideValidationError(errors)

    if not guides:
        raise ExcelGuideValidationError(
            ["No channel sheets found. Add at least one sheet whose name doesn't start with '_'."]
        )

    return guides


def _parse_sheet(sheet: object, sheet_name: str) -> ChannelGuide:
    rows_iter = sheet.iter_rows(values_only=True)  # type: ignore[attr-defined]
    try:
        header_raw = next(rows_iter)
    except StopIteration as err:
        raise _SheetValidationError([f"Sheet '{sheet_name}' is empty."]) from err

    header = [str(cell or "").strip().lower() for cell in header_raw]
    missing = [col for col in REQUIRED_COLUMNS if col not in header]
    if missing:
        raise _SheetValidationError(
            [f"Sheet '{sheet_name}' is missing required columns: {', '.join(missing)}."]
        )

    indices = {col: header.index(col) for col in EXPECTED_COLUMNS if col in header}

    sections: dict[str, list[GuidanceRow]] = {}
    section_order: list[str] = []
    errors: list[str] = []

    for row_index, raw_row in enumerate(rows_iter, start=2):
        if raw_row is None or all(cell is None or str(cell).strip() == "" for cell in raw_row):
            continue

        try:
            row_model = _row_to_model(raw_row, indices, sheet_name, row_index)
        except _SheetValidationError as exc:
            errors.extend(exc.messages)
            continue

        if row_model.section not in sections:
            sections[row_model.section] = []
            section_order.append(row_model.section)
        sections[row_model.section].append(row_model)

    if errors:
        raise _SheetValidationError(errors)

    if not section_order:
        raise _SheetValidationError(
            [f"Sheet '{sheet_name}' has a header but no data rows."]
        )

    return ChannelGuide(
        name=sheet_name,
        sections=[ChannelSection(heading=name, rows=sections[name]) for name in section_order],
    )


def _row_to_model(
    raw_row: tuple[object, ...],
    indices: dict[str, int],
    sheet_name: str,
    row_index: int,
) -> GuidanceRow:
    section = _cell(raw_row, indices.get("section"))
    question_type = _cell(raw_row, indices.get("question_type")).lower()
    guidance = _cell(raw_row, indices.get("guidance"))
    example = _cell(raw_row, indices.get("example")) or None
    required_raw = _cell(raw_row, indices.get("required")).lower()
    notes = _cell(raw_row, indices.get("notes")) or None

    issues: list[str] = []
    if not section:
        issues.append(f"{sheet_name}!A{row_index}: 'section' is empty.")
    if not guidance:
        issues.append(f"{sheet_name}!{row_index}: 'guidance' is empty.")
    try:
        qtype = QuestionType(question_type)
    except ValueError:
        issues.append(
            f"{sheet_name}!{row_index}: question_type '{question_type}' not in "
            f"{[t.value for t in QuestionType]}."
        )
        qtype = QuestionType.OPEN_TEXT  # placeholder, won't be used if issues raised

    if issues:
        raise _SheetValidationError(issues)

    return GuidanceRow(
        section=section,
        question_type=qtype,
        guidance=guidance,
        example=example,
        required=required_raw in TRUTHY,
        notes=notes,
    )


def _cell(row: tuple[object, ...], index: int | None) -> str:
    if index is None or index >= len(row):
        return ""
    value = row[index]
    if value is None:
        return ""
    return str(value).strip()


class _SheetValidationError(Exception):
    def __init__(self, messages: list[str]) -> None:
        self.messages = messages
        super().__init__("; ".join(messages))
