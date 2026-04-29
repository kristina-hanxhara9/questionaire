"""Parser for the multi-module questionnaire workbook.

Sheet naming convention (case-insensitive):

    core                      → ModuleKind.CORE       (84 channel-agnostic, industry-agnostic questions)
    industry__opticians       → ModuleKind.INDUSTRY   (74 industry-specific, channel-agnostic)
    industry__dentists        → ModuleKind.INDUSTRY
    unique__opticians__social_media → ModuleKind.UNIQUE (~6 industry × channel questions)
    unique__opticians__email
    _scratchpad               → skipped (single-leading-underscore = author notes)

Required columns per sheet (case-insensitive header in row 1):
    id, section, question_text, question_type
Optional columns:
    options, required, tags, notes

`options` is pipe-separated for single_choice / multi_choice / rating types.
`tags` is comma-separated topical tags used for dedup/alignment.
"""

from __future__ import annotations

import threading
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook import Workbook

from ..errors import (
    ConfigurationError,
    ExcelGuideValidationError,
    IndustryNotFoundError,
    ModuleNotFoundError,
)
from ..models import Module, ModuleKind, ModuleQuestion, ModuleSet, QuestionType

REQUIRED_COLUMNS = ("id", "section", "question_text", "question_type")
OPTIONAL_COLUMNS = ("options", "required", "tags", "notes")
ALL_COLUMNS = REQUIRED_COLUMNS + OPTIONAL_COLUMNS
TRUTHY = frozenset({"yes", "y", "true", "1", "t"})

CORE_SHEET = "core"
INDUSTRY_PREFIX = "industry__"
UNIQUE_PREFIX = "unique__"


_cache_lock = threading.Lock()
_cache: dict[Path, tuple[float, "_ParsedWorkbook"]] = {}


class _ParsedWorkbook:
    __slots__ = ("core", "industry_modules", "unique_modules")

    def __init__(
        self,
        core: Module | None,
        industry_modules: dict[str, Module],
        unique_modules: dict[tuple[str, str], Module],
    ) -> None:
        self.core = core
        self.industry_modules = industry_modules
        self.unique_modules = unique_modules


def is_module_workbook(path: Path) -> bool:
    """Lightweight detection: a workbook is module-style if it has a `core` sheet
    or any `industry__*` / `unique__*` sheets. Lets us coexist with the old
    channel-guidance workbook format without breaking it."""
    resolved = Path(path).resolve()
    if not resolved.exists():
        return False
    try:
        wb = load_workbook(filename=resolved, read_only=True, data_only=True)
    except Exception:
        return False
    try:
        names_lower = [n.lower() for n in wb.sheetnames]
        if CORE_SHEET in names_lower:
            return True
        if any(n.startswith(INDUSTRY_PREFIX) or n.startswith(UNIQUE_PREFIX) for n in names_lower):
            return True
        return False
    finally:
        wb.close()


def list_industries(path: Path) -> list[str]:
    return sorted(_load(path).industry_modules.keys())


def list_channels_for_industry(path: Path, industry: str) -> list[str]:
    parsed = _load(path)
    industry_norm = industry.strip().lower()
    if industry_norm not in parsed.industry_modules:
        raise IndustryNotFoundError(industry, list(parsed.industry_modules.keys()))
    return sorted(channel for ind, channel in parsed.unique_modules if ind == industry_norm)


def get_core_module(path: Path) -> Module:
    parsed = _load(path)
    if parsed.core is None:
        raise ModuleNotFoundError("core", "core", [])
    return parsed.core


def get_industry_module(path: Path, industry: str) -> Module:
    parsed = _load(path)
    industry_norm = industry.strip().lower()
    module = parsed.industry_modules.get(industry_norm)
    if module is None:
        raise IndustryNotFoundError(industry, list(parsed.industry_modules.keys()))
    return module


def get_unique_module(path: Path, industry: str, channel: str) -> Module:
    parsed = _load(path)
    industry_norm = industry.strip().lower()
    channel_norm = channel.strip().lower()
    module = parsed.unique_modules.get((industry_norm, channel_norm))
    if module is None:
        available = [
            f"{ind}/{ch}" for (ind, ch) in parsed.unique_modules if ind == industry_norm
        ]
        raise ModuleNotFoundError("unique", f"{industry}/{channel}", available)
    return module


def assemble_module_set(path: Path, industry: str, channel: str) -> ModuleSet:
    return ModuleSet(
        industry=industry.strip().lower(),
        channel=channel.strip().lower(),
        core=get_core_module(path),
        industry_standard=get_industry_module(path, industry),
        unique=get_unique_module(path, industry, channel),
    )


def clear_cache() -> None:
    with _cache_lock:
        _cache.clear()


def _load(path: Path) -> _ParsedWorkbook:
    resolved = Path(path).resolve()
    if not resolved.exists():
        raise ConfigurationError(f"Module workbook not found at {resolved}")

    mtime = resolved.stat().st_mtime
    with _cache_lock:
        cached = _cache.get(resolved)
        if cached is not None and cached[0] == mtime:
            return cached[1]

    wb = load_workbook(filename=resolved, read_only=True, data_only=True)
    try:
        parsed = _parse(wb)
    finally:
        wb.close()

    with _cache_lock:
        _cache[resolved] = (mtime, parsed)
    return parsed


def _parse(wb: Workbook) -> _ParsedWorkbook:
    core: Module | None = None
    industry_modules: dict[str, Module] = {}
    unique_modules: dict[tuple[str, str], Module] = {}
    errors: list[str] = []

    for raw_name in wb.sheetnames:
        name_lower = raw_name.lower()
        if name_lower.startswith("_") and not name_lower.startswith(INDUSTRY_PREFIX) and not name_lower.startswith(UNIQUE_PREFIX):
            continue

        try:
            if name_lower == CORE_SHEET:
                questions = _parse_questions(wb[raw_name], raw_name)
                core = Module(kind=ModuleKind.CORE, name="core", questions=questions)
            elif name_lower.startswith(INDUSTRY_PREFIX):
                industry = name_lower[len(INDUSTRY_PREFIX):]
                if not industry:
                    errors.append(f"Sheet '{raw_name}' has empty industry name after '{INDUSTRY_PREFIX}'.")
                    continue
                questions = _parse_questions(wb[raw_name], raw_name)
                industry_modules[industry] = Module(
                    kind=ModuleKind.INDUSTRY,
                    name=industry,
                    industry=industry,
                    questions=questions,
                )
            elif name_lower.startswith(UNIQUE_PREFIX):
                rest = name_lower[len(UNIQUE_PREFIX):]
                parts = rest.split("__", 1)
                if len(parts) != 2 or not parts[0] or not parts[1]:
                    errors.append(
                        f"Sheet '{raw_name}' must follow '{UNIQUE_PREFIX}<industry>__<channel>'."
                    )
                    continue
                industry, channel = parts[0], parts[1]
                questions = _parse_questions(wb[raw_name], raw_name)
                unique_modules[(industry, channel)] = Module(
                    kind=ModuleKind.UNIQUE,
                    name=f"{industry}:{channel}",
                    industry=industry,
                    channel=channel,
                    questions=questions,
                )
            # else: ignore unrecognized sheets so legacy workbooks coexist.
        except _SheetValidationError as exc:
            errors.extend(exc.messages)

    if errors:
        raise ExcelGuideValidationError(errors)

    return _ParsedWorkbook(core=core, industry_modules=industry_modules, unique_modules=unique_modules)


def _parse_questions(sheet: object, sheet_name: str) -> list[ModuleQuestion]:
    rows_iter = sheet.iter_rows(values_only=True)  # type: ignore[attr-defined]
    try:
        header_raw = next(rows_iter)
    except StopIteration as err:
        raise _SheetValidationError([f"Sheet '{sheet_name}' is empty."]) from err

    header = [str(cell or "").strip().lower() for cell in header_raw]
    missing = [col for col in REQUIRED_COLUMNS if col not in header]
    if missing:
        raise _SheetValidationError(
            [f"Sheet '{sheet_name}' is missing required columns: {', '.join(missing)}."]
        )

    indices = {col: header.index(col) for col in ALL_COLUMNS if col in header}

    questions: list[ModuleQuestion] = []
    seen_ids: set[str] = set()
    errors: list[str] = []

    for row_index, raw_row in enumerate(rows_iter, start=2):
        if raw_row is None or all(cell is None or str(cell).strip() == "" for cell in raw_row):
            continue

        qid = _cell(raw_row, indices.get("id"))
        section = _cell(raw_row, indices.get("section"))
        text = _cell(raw_row, indices.get("question_text"))
        qtype_raw = _cell(raw_row, indices.get("question_type")).lower()
        options_raw = _cell(raw_row, indices.get("options")) or None
        required_raw = _cell(raw_row, indices.get("required")).lower()
        tags_raw = _cell(raw_row, indices.get("tags")) or ""
        notes_raw = _cell(raw_row, indices.get("notes")) or None

        row_errors: list[str] = []
        if not qid:
            row_errors.append(f"{sheet_name}!{row_index}: 'id' is empty.")
        elif qid in seen_ids:
            row_errors.append(f"{sheet_name}!{row_index}: duplicate id '{qid}'.")
        if not section:
            row_errors.append(f"{sheet_name}!{row_index}: 'section' is empty.")
        if not text:
            row_errors.append(f"{sheet_name}!{row_index}: 'question_text' is empty.")
        try:
            qtype = QuestionType(qtype_raw)
        except ValueError:
            row_errors.append(
                f"{sheet_name}!{row_index}: question_type '{qtype_raw}' not in "
                f"{[t.value for t in QuestionType]}."
            )
            qtype = QuestionType.OPEN_TEXT

        if row_errors:
            errors.extend(row_errors)
            continue

        seen_ids.add(qid)
        questions.append(
            ModuleQuestion(
                id=qid,
                section=section,
                text=text,
                question_type=qtype,
                options=options_raw,
                required=required_raw in TRUTHY,
                tags=tags_raw,
                notes=notes_raw,
            )
        )

    if errors:
        raise _SheetValidationError(errors)
    if not questions:
        raise _SheetValidationError([f"Sheet '{sheet_name}' has a header but no data rows."])

    return questions


def _cell(row: tuple[object, ...], index: int | None) -> str:
    if index is None or index >= len(row):
        return ""
    value = row[index]
    if value is None:
        return ""
    return str(value).strip()


class _SheetValidationError(Exception):
    def __init__(self, messages: list[str]) -> None:
        self.messages = messages
        super().__init__("; ".join(messages))
